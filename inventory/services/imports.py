"""Import em massa de produtos (materiais) via CSV ou XLSX.

Faz upsert por SKU: linha com SKU já existente atualiza; SKU novo cria. Categoria
e fornecedor vêm por NOME e são criados se não existirem. Retorna um resumo com
criados/atualizados e erros por linha (nunca levanta para o chamador)."""
import csv
import io
import unicodedata

from ..extensions import db
from ..models.product import Product
from ..models.category import Category
from ..models.supplier import Supplier

# Sinônimos de cabeçalho (sem acento, minúsculo, espaços->_) -> campo canônico.
HEADER_ALIASES = {
    "sku": "sku", "codigo": "sku", "cod": "sku",
    "nome": "name", "name": "name", "produto": "name", "material": "name",
    "descricao": "description", "description": "description", "obs": "description",
    "categoria": "category", "category": "category",
    "fornecedor": "supplier", "supplier": "supplier",
    "estoque_minimo": "min_stock", "min_stock": "min_stock", "minimo": "min_stock",
    "preco": "price", "price": "price", "valor": "price", "custo": "price",
    "unidade": "unit", "unit": "unit", "un": "unit",
    "tipo": "item_type", "item_type": "item_type",
    "marca": "brand", "brand": "brand",
    "modelo": "model", "model": "model",
    "local": "location", "localizacao": "location", "location": "location",
    "patrimonio": "patrimony", "patrimony": "patrimony",
    "numero_serie": "serial_number", "serie": "serial_number", "serial_number": "serial_number",
}

# Colunas do modelo de exemplo (para download).
TEMPLATE_HEADERS = ["sku", "nome", "categoria", "fornecedor", "unidade",
                    "estoque_minimo", "preco", "marca", "modelo", "local"]

_SCALARS = ("description", "unit", "brand", "model", "location", "patrimony",
            "serial_number", "item_type")


# ===== Máquinas (computadores/notebooks/impressoras) =====
MACHINE_ALIASES = {
    "tipo": "kind", "kind": "kind",
    "nome": "name", "name": "name", "hostname": "name", "identificacao": "name",
    "marca": "brand", "brand": "brand",
    "modelo": "model", "model": "model",
    "responsavel": "assigned_user", "usuario": "assigned_user",
    "colaborador": "assigned_user", "assigned_user": "assigned_user",
    "ip": "ip_address", "endereco_ip": "ip_address", "ip_address": "ip_address",
    "setor": "sector", "sector": "sector", "departamento": "sector",
    "patrimonio": "patrimony", "patrimony": "patrimony",
    "numero_serie": "serial_number", "serie": "serial_number", "serial_number": "serial_number",
    "observacoes": "notes", "obs": "notes", "notes": "notes",
}

MACHINE_TEMPLATE_HEADERS = ["tipo", "nome", "marca", "modelo", "patrimonio",
                            "numero_serie", "ip", "setor", "responsavel"]

# Tipos aceitos na coluna "tipo" (normalizados).
_KINDS = {
    "computador": "computador", "pc": "computador", "desktop": "computador",
    "notebook": "notebook", "laptop": "notebook",
    "impressora": "impressora", "printer": "impressora",
}

_MACHINE_SCALARS = ("name", "brand", "model", "ip_address", "sector", "notes")


def _norm(h) -> str:
    t = unicodedata.normalize("NFKD", str(h or "")).encode("ascii", "ignore").decode()
    return t.strip().lower().replace(" ", "_")


def _map_header(h, aliases=None) -> str:
    return (aliases or HEADER_ALIASES).get(_norm(h), "")


def parse_table(file_storage, aliases=None) -> list:
    """Lê CSV ou XLSX e retorna uma lista de dicts com chaves canônicas."""
    fname = (file_storage.filename or "").lower()
    rows = []
    if fname.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(file_storage, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        cols = [_map_header(h, aliases) for h in (next(it, None) or [])]
        for r in it:
            rows.append({cols[i]: r[i] for i in range(min(len(cols), len(r))) if cols[i]})
    else:
        text = file_storage.read().decode("utf-8-sig", errors="replace")
        first = (text.splitlines() or [""])[0]
        try:
            dialect = csv.Sniffer().sniff(first, delimiters=",;\t")
        except Exception:  # noqa: BLE001
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        cols = [_map_header(h, aliases) for h in next(reader, [])]
        for r in reader:
            rows.append({cols[i]: (r[i] if i < len(r) else "")
                         for i in range(len(cols)) if cols[i]})
    return rows


def _as_int(v, default=0):
    try:
        return int(float(str(v).strip().replace(",", ".")))
    except (TypeError, ValueError):
        return default


def _as_float(v, default=0.0):
    try:
        return float(str(v).strip().replace(",", "."))
    except (TypeError, ValueError):
        return default


def _resolve(model, name, created_counter):
    nome = str(name or "").strip()
    if not nome:
        return None
    o = model.query.filter(db.func.lower(model.name) == nome.lower()).first()
    if o:
        return o.id
    o = model(name=nome)
    db.session.add(o)
    db.session.flush()
    created_counter[0] += 1
    return o.id


def import_products(rows) -> dict:
    """Upsert por SKU. Retorna resumo {created, updated, errors[], categorias,
    fornecedores}."""
    summary = {"created": 0, "updated": 0, "errors": [], "categorias": 0, "fornecedores": 0}
    cat_new, sup_new = [0], [0]

    for i, row in enumerate(rows, start=2):   # linha 1 = cabeçalho
        name = str(row.get("name") or "").strip()
        sku = str(row.get("sku") or "").strip()
        if not name and not sku:
            continue  # linha em branco
        if not sku:
            summary["errors"].append(f"Linha {i}: SKU obrigatório (ignorada)")
            continue
        if not name:
            summary["errors"].append(f"Linha {i}: nome obrigatório (ignorada)")
            continue

        try:
            cat_id = _resolve(Category, row.get("category"), cat_new)
            sup_id = _resolve(Supplier, row.get("supplier"), sup_new)
            values = {
                "name": name,
                "category_id": cat_id,
                "supplier_id": sup_id,
                "min_stock": _as_int(row.get("min_stock")),
                "price": _as_float(row.get("price")),
            }
            for f in _SCALARS:
                if f in row and str(row.get(f) or "").strip():
                    values[f] = str(row.get(f)).strip()
            values.setdefault("unit", values.get("unit") or "UN")
            values.setdefault("item_type", values.get("item_type") or "product")

            p = Product.query.filter(db.func.lower(Product.sku) == sku.lower()).first()
            if p:
                for k, v in values.items():
                    setattr(p, k, v)
                summary["updated"] += 1
            else:
                db.session.add(Product(sku=sku, **values))
                summary["created"] += 1
        except Exception as e:  # noqa: BLE001
            db.session.rollback()
            summary["errors"].append(f"Linha {i} (SKU {sku}): {e}")
            continue

    try:
        db.session.commit()
    except Exception as e:  # noqa: BLE001
        db.session.rollback()
        summary["errors"].append(f"Falha ao gravar: {e}")
    summary["categorias"] = cat_new[0]
    summary["fornecedores"] = sup_new[0]
    return summary


def import_machines(rows) -> dict:
    """Upsert de máquinas por PATRIMÔNIO (ou nº de série). Sem nenhum dos dois,
    cria um registro novo. Retorna {created, updated, errors[]}."""
    from ..models.machine import Machine
    from . import people

    summary = {"created": 0, "updated": 0, "errors": []}
    for i, row in enumerate(rows, start=2):   # linha 1 = cabeçalho
        nome = str(row.get("name") or "").strip()
        pat = str(row.get("patrimony") or "").strip()
        serie = str(row.get("serial_number") or "").strip()
        modelo = str(row.get("model") or "").strip()
        if not any((nome, pat, serie, modelo)):
            continue  # linha em branco
        if not (nome or modelo):
            summary["errors"].append(f"Linha {i}: informe ao menos 'nome' ou 'modelo' (ignorada)")
            continue

        try:
            kind = _KINDS.get(_norm(row.get("kind")), "computador")
            values = {"kind": kind}
            for f in _MACHINE_SCALARS:
                if f in row and str(row.get(f) or "").strip():
                    values[f] = str(row.get(f)).strip()
            resp = str(row.get("assigned_user") or "").strip()
            if resp:
                values["assigned_user"] = resp
                # Mantém a FK em dia (mesma regra do cadastro pela tela).
                values["user_id"] = people.user_id_for(resp)
                if not values.get("sector"):
                    values["sector"] = people.sector_for(resp) or None
            if pat:
                values["patrimony"] = pat
            if serie:
                values["serial_number"] = serie

            # Chave de upsert: patrimônio > nº de série > (nenhuma: cria)
            m = None
            if pat:
                m = Machine.query.filter(db.func.lower(Machine.patrimony) == pat.lower()).first()
            if m is None and serie:
                m = Machine.query.filter(
                    db.func.lower(Machine.serial_number) == serie.lower()).first()

            if m:
                for k, v in values.items():
                    setattr(m, k, v)
                summary["updated"] += 1
            else:
                db.session.add(Machine(**values))
                summary["created"] += 1
        except Exception as e:  # noqa: BLE001
            db.session.rollback()
            summary["errors"].append(f"Linha {i}: {e}")
            continue

    try:
        db.session.commit()
    except Exception as e:  # noqa: BLE001
        db.session.rollback()
        summary["errors"].append(f"Falha ao gravar: {e}")
    return summary
