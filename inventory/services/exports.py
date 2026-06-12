"""
Geração de planilhas Excel (.xlsx) para exportação.

Uso:
    from ..services.exports import xlsx_response
    return xlsx_response("Chamados", ["Código", "Título"], rows, filename="chamados")
"""
import io

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def build_xlsx(sheet_title: str, headers: list, rows: list) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Planilha")[:31]

    head_fill = PatternFill("solid", fgColor="2E8B57")
    head_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)

    # Largura automática (aproximada) por coluna
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        maxlen = len(str(headers[col - 1]))
        for row in rows:
            if col - 1 < len(row) and row[col - 1] is not None:
                maxlen = max(maxlen, len(str(row[col - 1])))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 60)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def xlsx_response(sheet_title: str, headers: list, rows: list, filename: str = "export"):
    buf = build_xlsx(sheet_title, headers, rows)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename}.xlsx",
    )
